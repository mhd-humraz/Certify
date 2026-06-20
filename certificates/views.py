import io
import os
import zipfile
import csv
from datetime import date

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Count
from django.core.mail import EmailMessage
from django.utils import timezone

from .models import Event, CertificateTemplate, Participant, Certificate
from .forms import (EventForm, TemplateForm, ParticipantForm,
                    ExcelUploadForm, BulkGenerateForm)
from .utils import generate_certificate_image, image_to_pdf

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ─── Auth ─────────────────────────────────────────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    from django.contrib.auth.views import LoginView
    return LoginView.as_view(template_name='registration/login.html')(request)


@login_required
def logout_view(request):
    logout(request)
    return redirect('login')


# ─── Dashboard ────────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    ctx = {
        'total_events': Event.objects.count(),
        'total_certificates': Certificate.objects.count(),
        'total_emails_sent': Certificate.objects.filter(email_sent=True).count(),
        'active_templates': CertificateTemplate.objects.filter(is_active=True).count(),
        'recent_events': Event.objects.order_by('-created_at')[:5],
        'recent_certificates': Certificate.objects.select_related(
            'participant', 'event').order_by('-created_at')[:8],
    }
    return render(request, 'certificates/dashboard.html', ctx)


# ─── Events ───────────────────────────────────────────────────────────────────

@login_required
def event_list(request):
    events = Event.objects.annotate(cert_count=Count('certificate')).order_by('-created_at')
    return render(request, 'certificates/event_list.html', {'events': events})


@login_required
def event_create(request):
    form = EventForm(request.POST or None)
    if form.is_valid():
        event = form.save(commit=False)
        event.created_by = request.user
        event.save()
        messages.success(request, f'Event "{event.event_name}" created successfully.')
        return redirect('event_list')
    return render(request, 'certificates/event_form.html', {'form': form, 'title': 'Create Event'})


@login_required
def event_edit(request, pk):
    event = get_object_or_404(Event, pk=pk)
    form = EventForm(request.POST or None, instance=event)
    if form.is_valid():
        form.save()
        messages.success(request, 'Event updated.')
        return redirect('event_list')
    return render(request, 'certificates/event_form.html', {'form': form, 'title': 'Edit Event'})


@login_required
def event_delete(request, pk):
    event = get_object_or_404(Event, pk=pk)
    if request.method == 'POST':
        name = event.event_name
        event.delete()
        messages.success(request, f'Event "{name}" deleted.')
        return redirect('event_list')
    return render(request, 'certificates/confirm_delete.html', {'object': event, 'type': 'Event'})


@login_required
def event_detail(request, pk):
    event = get_object_or_404(Event, pk=pk)
    participants = Participant.objects.filter(event=event)
    certificates = Certificate.objects.filter(event=event).select_related('participant')
    return render(request, 'certificates/event_detail.html', {
        'event': event,
        'participants': participants,
        'certificates': certificates,
    })


# ─── Participants ─────────────────────────────────────────────────────────────

@login_required
def participant_list(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    participants = Participant.objects.filter(event=event)
    form = ParticipantForm()

    if request.method == 'POST':
        form = ParticipantForm(request.POST)
        if form.is_valid():
            p = form.save(commit=False)
            p.event = event
            p.save()
            messages.success(request, f'{p.name} added.')
            return redirect('participant_list', event_pk=event_pk)

    return render(request, 'certificates/participant_list.html', {
        'event': event, 'participants': participants, 'form': form
    })


@login_required
def participant_delete(request, pk):
    p = get_object_or_404(Participant, pk=pk)
    event_pk = p.event_id
    p.delete()
    messages.success(request, 'Participant removed.')
    return redirect('participant_list', event_pk=event_pk)


@login_required
def upload_participants(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    form = ExcelUploadForm()
    errors = []
    added = 0

    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            f = request.FILES['excel_file']
            fname = f.name.lower()

            rows = []
            if fname.endswith('.csv'):
                decoded = f.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded))
                for row in reader:
                    rows.append(row)
            elif fname.endswith('.xlsx') and HAS_OPENPYXL:
                wb = openpyxl.load_workbook(f)
                ws = wb.active
                headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append({headers[i]: (str(v).strip() if v else '') for i, v in enumerate(row)})
            else:
                errors.append('Unsupported file format. Please upload .xlsx or .csv')

            for i, row in enumerate(rows, start=2):
                # Normalize keys
                row = {k.lower().strip(): v for k, v in row.items() if k}
                name = row.get('name', '').strip()
                email = row.get('email', '').strip()
                college = row.get('college', '').strip()
                course = row.get('course', '').strip()

                if not name or not email:
                    errors.append(f'Row {i}: Missing name or email.')
                    continue

                obj, created = Participant.objects.get_or_create(
                    event=event, email=email,
                    defaults={'name': name, 'college': college, 'course': course}
                )
                if created:
                    added += 1

            if added:
                messages.success(request, f'{added} participant(s) imported.')
            if errors:
                for e in errors[:5]:
                    messages.warning(request, e)

            return redirect('participant_list', event_pk=event_pk)

    return render(request, 'certificates/upload_participants.html', {
        'event': event, 'form': form
    })


# ─── Templates ────────────────────────────────────────────────────────────────

@login_required
def template_list(request):
    templates = CertificateTemplate.objects.order_by('-created_at')
    return render(request, 'certificates/template_list.html', {'templates': templates})


@login_required
def template_create(request):
    form = TemplateForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        t = form.save(commit=False)
        t.created_by = request.user
        t.save()
        messages.success(request, f'Template "{t.template_name}" uploaded.')
        return redirect('template_list')
    return render(request, 'certificates/template_form.html', {'form': form, 'title': 'Upload Template'})


@login_required
def template_delete(request, pk):
    t = get_object_or_404(CertificateTemplate, pk=pk)
    if request.method == 'POST':
        t.delete()
        messages.success(request, 'Template deleted.')
        return redirect('template_list')
    return render(request, 'certificates/confirm_delete.html', {'object': t, 'type': 'Template'})


# ─── Certificate Generation ───────────────────────────────────────────────────

@login_required
def generate_bulk(request):
    form = BulkGenerateForm(request.POST or None)
    if form.is_valid():
        event = form.cleaned_data['event']
        template_obj = form.cleaned_data['template']
        base_url = form.cleaned_data.get('base_url', 'http://localhost:8000').rstrip('/')

        participants = Participant.objects.filter(event=event)
        if not participants.exists():
            messages.error(request, 'No participants found for this event.')
            return redirect('generate_bulk')

        generated = 0
        skipped = 0
        for participant in participants:
            if Certificate.objects.filter(participant=participant, event=event).exists():
                skipped += 1
                continue

            cert_id = Certificate.generate_id()
            verify_url = f"{base_url}/certificate/verify/{cert_id}/"

            # Generate image and PDF
            cert_img = generate_certificate_image(
                template_obj=template_obj,
                participant_name=participant.name,
                event_name=event.event_name,
                issue_date=date.today(),
                certificate_id=cert_id,
                verify_url=verify_url,
                certificate_type=event.certificate_type,
            )
            pdf_bytes = image_to_pdf(cert_img)

            # Save PDF
            from django.core.files.base import ContentFile
            cert = Certificate(
                certificate_id=cert_id,
                participant=participant,
                event=event,
                template=template_obj,
                issue_date=date.today(),
            )
            cert.pdf_file.save(f'{cert_id}.pdf', ContentFile(pdf_bytes), save=False)
            cert.save()
            generated += 1

        messages.success(request,
            f'{generated} certificate(s) generated. {skipped} already existed (skipped).')
        return redirect('event_detail', pk=event.pk)

    return render(request, 'certificates/generate_bulk.html', {'form': form})


@login_required
def certificate_download(request, pk):
    cert = get_object_or_404(Certificate, pk=pk)
    if cert.pdf_file and os.path.exists(cert.pdf_file.path):
        with open(cert.pdf_file.path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{cert.certificate_id}.pdf"'
            return response
    messages.error(request, 'Certificate file not found.')
    return redirect('dashboard')


@login_required
def download_zip(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    certificates = Certificate.objects.filter(event=event)

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for cert in certificates:
            if cert.pdf_file and os.path.exists(cert.pdf_file.path):
                zf.write(cert.pdf_file.path,
                         arcname=f'{cert.participant.name}_{cert.certificate_id}.pdf')

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{event.event_name}_certificates.zip"'
    return response


# ─── Email ────────────────────────────────────────────────────────────────────

@login_required
def send_certificate_email(request, cert_pk):
    cert = get_object_or_404(Certificate, pk=cert_pk)
    if not cert.pdf_file:
        messages.error(request, 'No PDF to send.')
        return redirect('event_detail', pk=cert.event_id)

    try:
        email = EmailMessage(
            subject='Your Certificate is Ready – CertifyPro',
            body=(
                f'Dear {cert.participant.name},\n\n'
                f'Congratulations! Please find your certificate for '
                f'"{cert.event.event_name}" attached.\n\n'
                f'Certificate ID: {cert.certificate_id}\n'
                f'Verify at: /certificate/verify/{cert.certificate_id}/\n\n'
                f'Best regards,\nCertifyPro Team'
            ),
            to=[cert.participant.email],
        )
        with open(cert.pdf_file.path, 'rb') as f:
            email.attach(f'{cert.certificate_id}.pdf', f.read(), 'application/pdf')
        email.send()

        cert.email_sent = True
        cert.email_sent_at = timezone.now()
        cert.save()
        messages.success(request, f'Certificate emailed to {cert.participant.email}.')
    except Exception as e:
        messages.error(request, f'Email failed: {e}')

    return redirect('event_detail', pk=cert.event_id)


@login_required
def send_bulk_email(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    certificates = Certificate.objects.filter(event=event, email_sent=False)
    sent = 0
    for cert in certificates:
        if not cert.pdf_file or not os.path.exists(cert.pdf_file.path):
            continue
        try:
            email = EmailMessage(
                subject='Your Certificate is Ready – CertifyPro',
                body=(
                    f'Dear {cert.participant.name},\n\n'
                    f'Congratulations! Please find your certificate for '
                    f'"{event.event_name}" attached.\n\n'
                    f'Certificate ID: {cert.certificate_id}\n\n'
                    f'Best regards,\nCertifyPro Team'
                ),
                to=[cert.participant.email],
            )
            with open(cert.pdf_file.path, 'rb') as f:
                email.attach(f'{cert.certificate_id}.pdf', f.read(), 'application/pdf')
            email.send()
            cert.email_sent = True
            cert.email_sent_at = timezone.now()
            cert.save()
            sent += 1
        except Exception:
            pass

    messages.success(request, f'{sent} email(s) sent.')
    return redirect('event_detail', pk=event_pk)


# ─── Verification ─────────────────────────────────────────────────────────────

def verify_certificate(request, certificate_id):
    cert = Certificate.objects.filter(certificate_id=certificate_id).first()
    return render(request, 'certificates/verify.html', {
        'cert': cert,
        'certificate_id': certificate_id,
    })


def verify_search(request):
    cert = None
    searched = False
    cert_id = request.GET.get('q', '').strip()
    if cert_id:
        searched = True
        cert = Certificate.objects.filter(certificate_id=cert_id).first()
    return render(request, 'certificates/verify_search.html', {
        'cert': cert, 'searched': searched, 'cert_id': cert_id
    })


# ─── Reports ──────────────────────────────────────────────────────────────────

@login_required
def reports(request):
    from django.db.models.functions import TruncDate, TruncMonth
    daily = (Certificate.objects
             .annotate(day=TruncDate('created_at'))
             .values('day')
             .annotate(count=Count('id'))
             .order_by('-day')[:30])

    monthly = (Certificate.objects
               .annotate(month=TruncMonth('created_at'))
               .values('month')
               .annotate(count=Count('id'))
               .order_by('-month')[:12])

    email_rate = 0
    total_certs = Certificate.objects.count()
    if total_certs:
        email_rate = round(Certificate.objects.filter(email_sent=True).count() / total_certs * 100, 1)

    ctx = {
        'daily': list(daily),
        'monthly': list(monthly),
        'email_rate': email_rate,
        'total_certs': total_certs,
        'total_events': Event.objects.count(),
        'emails_sent': Certificate.objects.filter(email_sent=True).count(),
        'event_stats': Event.objects.annotate(cert_count=Count('certificate')).order_by('-cert_count')[:10],
    }
    return render(request, 'certificates/reports.html', ctx)
